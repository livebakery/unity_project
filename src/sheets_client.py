"""Read raw rows from a Google Sheet OR an .xlsx file in Drive.

The mentor's file (`Valuation.xlsx`) is uploaded to Drive in Excel format,
so the Sheets API v4 will refuse it. We detect mimeType via Drive API and
fall back to downloading + parsing with openpyxl.
"""
from __future__ import annotations

import io
import json
import os
from typing import Optional

import gspread
import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GSHEET_MIMETYPE = "application/vnd.google-apps.spreadsheet"


def _credentials() -> Credentials:
    raw = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not raw:
        raise RuntimeError(
            "GOOGLE_SERVICE_ACCOUNT_JSON env var is not set. "
            "Paste the service account JSON content into this variable."
        )
    info = json.loads(raw)
    return Credentials.from_service_account_info(info, scopes=SCOPES)


def _drive_service():
    return build("drive", "v3", credentials=_credentials(), cache_discovery=False)


def _read_xlsx(file_id: str, worksheet_name: Optional[str]) -> list[list[str]]:
    """Download an Excel file from Drive and return all rows as strings."""
    svc = _drive_service()
    request = svc.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _status, done = downloader.next_chunk()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    try:
        if worksheet_name:
            ws = wb[worksheet_name]
        else:
            ws = _pick_best_worksheet(wb)
        rows: list[list[str]] = []
        for row in ws.iter_rows():
            rows.append([_cell_to_str(c) for c in row])
        return rows
    finally:
        wb.close()


def _pick_best_worksheet(wb: openpyxl.Workbook):
    """Pick the worksheet whose first non-empty row contains the keyword 'Stock'.

    Falls back to the first worksheet.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and cell.strip().lower() == "stock":
                    return ws
    return wb.worksheets[0]


def _cell_to_str(cell) -> str:
    """Render an openpyxl cell to a string the parser can consume.

    Percent-formatted cells are stored as decimals in Excel (14.72% → 0.1472),
    so re-multiply by 100 and append "%" to match the human-facing value.
    """
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or ""
    if isinstance(v, (int, float)) and "%" in fmt:
        return f"{v * 100:g}%"
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return f"{v:g}"
    return str(v)


def _read_gsheet(file_id: str, worksheet_name: Optional[str]) -> list[list[str]]:
    client = gspread.authorize(_credentials())
    spreadsheet = client.open_by_key(file_id)
    ws = (
        spreadsheet.worksheet(worksheet_name)
        if worksheet_name
        else spreadsheet.sheet1
    )
    return ws.get_all_values()


def read_sheet(sheet_id: str, worksheet_name: Optional[str] = None) -> list[list[str]]:
    """Return all values as a list of rows.

    Handles both native Google Sheets and Excel (.xlsx) files in Drive.
    """
    svc = _drive_service()
    meta = svc.files().get(fileId=sheet_id, fields="mimeType,name").execute()
    if meta.get("mimeType") == XLSX_MIMETYPE:
        return _read_xlsx(sheet_id, worksheet_name)
    return _read_gsheet(sheet_id, worksheet_name)
